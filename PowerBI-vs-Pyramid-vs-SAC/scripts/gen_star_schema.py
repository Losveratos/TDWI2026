#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Baut aus der bestehenden ZlatyLev_Sales_Report_1.xlsx (echte Monatsvolumina hl
je Produkt, 2023-2026) ein sauberes Sternschema:

  ZlatyLev_DWH.xlsx   -> DimDate, DimProduct, DimCustomer, DimKennzahl, FactActual
  ZlatyLev_Plan.xlsx  -> Plan (Kreuztabelle Produkt x Monat, Net Sales EUR, je Jahr)
  dwh_csv/*.csv       -> jede Tabelle als CSV (fuer DB-Load)

Deterministisch (fester Seed). Re-run ueberschreibt die Ausgaben.
"""
import csv
import os
import re
import datetime as dt
from collections import defaultdict

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SRC = os.path.join(BASE, "ZlatyLev_Sales_Report_1.xlsx")
DWH_XLSX = os.path.join(BASE, "ZlatyLev_DWH.xlsx")
PLAN_XLSX = os.path.join(BASE, "ZlatyLev_Plan.xlsx")
CSV_DIR = os.path.join(BASE, "dwh_csv")

YEARS = [2023, 2024, 2025, 2026]
MONTHS_DE = ["Jan", "Feb", "Mär", "Apr", "Mai", "Jun",
             "Jul", "Aug", "Sep", "Okt", "Nov", "Dez"]

HDR_FILL = PatternFill("solid", fgColor="0F1E2E")
HDR_FONT = Font(bold=True, color="FFFFFF", name="Calibri")

# ---------------------------------------------------------------------------
# Stammdaten
# ---------------------------------------------------------------------------

def norm(s):
    return re.sub(r"\s+", " ", str(s)).strip()

# Produkt-Stammdaten: norm-Name aus Quelle -> Attribute
# netprice / unitcost = 2023-Basis in EUR je hl (netto realisiert bzw. Stueckkost)
PRODUCTS = [
    # src_match, ProductName, Category, Style, ABV, Launch, netprice, unitcost
    ("Zlatý Lev 10° — Desítka",        "Zlatý Lev 10° Desítka",        "Pale Lager",     "Pale Lager",        4.1, 1995, 138, 82),
    ("Zlatý Lev 12° — Dvanáctka",      "Zlatý Lev 12° Dvanáctka",      "Pale Lager",     "Premium Pale Lager",5.0, 1995, 152, 86),
    ("Praha Premium 13°",              "Praha Premium 13°",            "Pale Lager",     "Strong Pale Lager", 5.5, 2008, 168, 92),
    ("Černý Lev 11°",                  "Černý Lev 11°",                "Dark Lager",     "Dark Lager",        4.6, 2001, 150, 88),
    ("Vyšehrad Tmavý 13°",             "Vyšehrad Tmavý 13°",           "Dark Lager",     "Strong Dark Lager", 5.2, 2011, 172, 95),
    ("Karlův IPA",                     "Karlův IPA",                   "Specialty",      "India Pale Ale",    6.2, 2018, 205, 118),
    ("Vltava Pšenice (wheat)",         "Vltava Pšenice",               "Specialty",      "Wheat Beer",        5.1, 2016, 192, 110),
    ("Anniversary 1882 (limited ed.)", "Anniversary 1882",             "Specialty",      "Limited Edition",   6.5, 2024, 240, 132),
    ("Smíchov Hazy IPA",               "Smíchov Hazy IPA",             "Specialty",      "Hazy / NE IPA",     6.0, 2025, 210, 120),
    ("Lev Free",                       "Lev Free",                     "Non-Alcoholic",  "Alcohol-free Lager",0.4, 2014, 132, 80),
    ("Lev Citrón Radler",              "Lev Citrón Radler",            "Non-Alcoholic",  "Radler / Shandy",   2.0, 2017, 124, 76),
]
PRODUCT_KEY = {p[1]: i + 1 for i, p in enumerate(PRODUCTS)}  # ProductName -> key

# Grosskunden mit Anschrift
CUSTOMERS = [
    # Name, Type, Street, PostalCode, City, Region, Country, Since, KAM, base_weight
    ("Albert Česká republika s.r.o.", "Retail Chain",  "Radlická 117",          "158 00", "Praha",      "CZ-Prag",    "Czechia",  2006, "J. Nováková",  0.14),
    ("BILLA spol. s r.o.",            "Retail Chain",  "Modletice 67",          "251 01", "Říčany",     "CZ-Mitte",   "Czechia",  2007, "J. Nováková",  0.10),
    ("Kaufland Česká republika v.o.s.","Retail Chain", "Bělohorská 2428",       "169 00", "Praha",      "CZ-Prag",    "Czechia",  2008, "M. Dvořák",    0.12),
    ("Tesco Stores ČR a.s.",          "Retail Chain",  "Vršovická 1527",        "100 00", "Praha",      "CZ-Prag",    "Czechia",  2005, "M. Dvořák",    0.10),
    ("Lidl Česká republika s.r.o.",   "Retail Chain",  "Nárožní 1359",          "158 00", "Praha",      "CZ-Prag",    "Czechia",  2010, "M. Dvořák",    0.11),
    ("MAKRO Cash & Carry ČR s.r.o.",  "Wholesale",     "Jeremiášova 1249",      "155 00", "Praha",      "CZ-Prag",    "Czechia",  2004, "P. Horák",     0.09),
    ("PramenPlus Distribuce s.r.o.",  "Wholesale",     "Tuřanka 1222",          "627 00", "Brno",       "CZ-Mähren",  "Czechia",  2009, "P. Horák",     0.07),
    ("Ambiente (Lokál) Restaurace",   "On-Trade",      "Dlouhá 731",            "110 00", "Praha",      "CZ-Prag",    "Czechia",  2012, "L. Marešová",  0.05),
    ("Kolkovna Group s.r.o.",         "On-Trade",      "V Kolkovně 910",        "110 00", "Praha",      "CZ-Prag",    "Czechia",  2013, "L. Marešová",  0.04),
    ("Getränke Hoffmann GmbH",        "Export",        "Großenhainer Str. 9",   "01097",  "Dresden",    "DE-Sachsen", "Germany",  2015, "S. Berger",    0.08),
    ("TESCO Stores SR a.s.",          "Export",        "Kamenné nám. 1/A",      "815 61", "Bratislava", "SK-West",    "Slovakia", 2011, "S. Berger",    0.06),
    ("Żabka Polska sp. z o.o.",       "Export",        "ul. Stanisława Matyi 8","61-586", "Poznań",     "PL-West",    "Poland",   2017, "S. Berger",    0.04),
]

# Kategorie-Affinitaet je Kundentyp (relative Gewichtung)
CAT_MULT = {
    "Retail Chain": {"Pale Lager": 1.2, "Dark Lager": 1.0, "Specialty": 0.7, "Non-Alcoholic": 1.3},
    "Wholesale":    {"Pale Lager": 1.1, "Dark Lager": 1.0, "Specialty": 1.0, "Non-Alcoholic": 0.9},
    "On-Trade":     {"Pale Lager": 1.0, "Dark Lager": 1.1, "Specialty": 1.5, "Non-Alcoholic": 0.6},
    "Export":       {"Pale Lager": 1.3, "Dark Lager": 0.9, "Specialty": 1.1, "Non-Alcoholic": 0.8},
}
DISC_RATE = {"Retail Chain": 0.12, "Wholesale": 0.18, "On-Trade": 0.08, "Export": 0.15}

# Kennzahlen (Account-Dimension) – Einheit liegt auf dem Account (SAC/Pyramid-Stil)
ACCOUNTS = [
    # key, name, group, unit, sort
    (1, "Volume",    "Operativ", "hl",  10),
    (2, "Net Sales", "GuV",      "EUR", 20),
    (3, "Discounts", "GuV",      "EUR", 30),
    (4, "COGS",      "GuV",      "EUR", 40),
]

PRICE_GROWTH = 0.03   # netprice +3 %/Jahr
COST_GROWTH = 0.025   # unitcost +2.5 %/Jahr

# ---------------------------------------------------------------------------
# 1) Echte Volumina aus der Quell-Excel lesen
# ---------------------------------------------------------------------------

def read_source_volumes():
    """-> vol[(ProductName, year, month)] = hl   (month 1..12)"""
    wb = openpyxl.load_workbook(SRC, data_only=True)
    match = {norm(p[0]): p[1] for p in PRODUCTS}
    vol = {}
    for year in YEARS:
        ws = wb[f"Sales {year}"]
        for row in ws.iter_rows(values_only=True):
            label = norm(row[0]) if row[0] else ""
            label = norm(label.replace("(NEW)", "").replace("(new)", ""))
            if label in match:
                pname = match[label]
                for m in range(1, 13):
                    v = row[m]
                    vol[(pname, year, m)] = int(v) if isinstance(v, (int, float)) else 0
    return vol

# ---------------------------------------------------------------------------
# 2) FactActual long-format aufbauen
# ---------------------------------------------------------------------------

def build_fact_actual(vol):
    """Liefert Liste von Dicts: Date, CustomerKey, ProductKey, AccountKey, Value"""
    rows = []
    cat_of = {p[1]: p[2] for p in PRODUCTS}
    price0 = {p[1]: p[6] for p in PRODUCTS}
    cost0 = {p[1]: p[7] for p in PRODUCTS}

    for (pname, year, month), total in vol.items():
        if total <= 0:
            continue
        cat = cat_of[pname]
        pkey = PRODUCT_KEY[pname]
        date = dt.date(year, month, 1)
        yidx = year - 2023
        price = price0[pname] * (1 + PRICE_GROWTH) ** yidx
        cost = cost0[pname] * (1 + COST_GROWTH) ** yidx

        # Gewichte je Kunde fuer dieses Produkt
        weights = []
        for ci, c in enumerate(CUSTOMERS):
            ctype = c[1]
            w = c[9] * CAT_MULT[ctype][cat]
            weights.append(w)
        wsum = sum(weights)

        # Volumen auf Kunden verteilen, ganzzahlig, Rest auf groessten Kunden
        alloc = [int(round(total * w / wsum)) for w in weights]
        diff = total - sum(alloc)
        if diff != 0:
            j = max(range(len(alloc)), key=lambda i: weights[i])
            alloc[j] += diff

        for ci, c in enumerate(CUSTOMERS):
            v_hl = alloc[ci]
            if v_hl <= 0:
                continue
            ckey = ci + 1
            net = v_hl * price
            disc = net * DISC_RATE[c[1]] / (1 - DISC_RATE[c[1]])  # Rabatt auf Brutto
            cogs = v_hl * cost
            rows.append({"Date": date, "CustomerKey": ckey, "ProductKey": pkey,
                         "AccountKey": 1, "Value": v_hl})
            rows.append({"Date": date, "CustomerKey": ckey, "ProductKey": pkey,
                         "AccountKey": 2, "Value": round(net)})
            rows.append({"Date": date, "CustomerKey": ckey, "ProductKey": pkey,
                         "AccountKey": 3, "Value": round(disc)})
            rows.append({"Date": date, "CustomerKey": ckey, "ProductKey": pkey,
                         "AccountKey": 4, "Value": round(cogs)})
    rows.sort(key=lambda r: (r["Date"], r["ProductKey"], r["CustomerKey"], r["AccountKey"]))
    return rows

# ---------------------------------------------------------------------------
# 3) Plan-Kreuztabelle (Net Sales EUR) je Produkt x Monat x Jahr
# ---------------------------------------------------------------------------

def build_plan(fact_rows):
    """Plan top-down: Vorjahres-Ist-Umsatz x Zielwachstum, verteilt ueber
       gemitteltes Saisonprofil. -> plan[(ProductName, year)] = [12 Monatswerte]"""
    # actual net sales je (product, year, month)
    net = defaultdict(float)
    pname_of = {v: k for k, v in PRODUCT_KEY.items()}
    for r in fact_rows:
        if r["AccountKey"] == 2:
            net[(pname_of[r["ProductKey"]], r["Date"].year, r["Date"].month)] += r["Value"]

    annual = defaultdict(float)
    for (p, y, m), v in net.items():
        annual[(p, y)] += v

    # Saisonprofil je Produkt (Mittel ueber Jahre mit Umsatz)
    profile = {}
    for p in PRODUCT_KEY:
        shares = [0.0] * 12
        cnt = 0
        for y in YEARS:
            if annual[(p, y)] > 0:
                cnt += 1
                for m in range(1, 13):
                    shares[m - 1] += net[(p, y, m)] / annual[(p, y)]
        if cnt:
            shares = [s / cnt for s in shares]
        else:
            shares = [1 / 12] * 12
        profile[p] = shares

    growth = {2023: 0.05, 2024: 0.08, 2025: 0.07, 2026: 0.05}
    plan = {}
    for p in PRODUCT_KEY:
        for y in YEARS:
            prev = annual[(p, y - 1)]
            base = prev if prev > 0 else annual[(p, y)]  # Fallback (z.B. Neu-Launch)
            target = base * (1 + growth[y])
            vals = [round(target * profile[p][m] / 100) * 100 for m in range(12)]
            plan[(p, y)] = vals
    return plan

# ---------------------------------------------------------------------------
# 4) DimDate
# ---------------------------------------------------------------------------

def build_dim_date():
    rows = []
    d = dt.date(2023, 1, 1)
    end = dt.date(2026, 12, 31)
    daynames = ["Mo", "Di", "Mi", "Do", "Fr", "Sa", "So"]
    while d <= end:
        m = d.month
        q = (m - 1) // 3 + 1
        rows.append({
            "Date": d,
            "DateKey": d.year * 10000 + m * 100 + d.day,
            "Year": d.year,
            "Quarter": q,
            "QuarterName": f"Q{q}",
            "MonthNo": m,
            "MonthName": MONTHS_DE[m - 1],
            "MonthYear": f"{MONTHS_DE[m - 1]} {d.year}",
            "MonthYearSort": d.year * 100 + m,
            "MonthStart": dt.date(d.year, m, 1),
            "DayOfMonth": d.day,
            "DayName": daynames[d.weekday()],
        })
        d += dt.timedelta(days=1)
    return rows

# ---------------------------------------------------------------------------
# Excel-Helfer: Sheet als saubere Tabelle (ListObject) schreiben
# ---------------------------------------------------------------------------

def write_table(ws, headers, rows, table_name, date_cols=(), int_cols=()):
    ws.append(headers)
    for r in rows:
        ws.append([r.get(h) for h in headers])
    nrows = len(rows) + 1
    ncols = len(headers)
    # Header-Styling
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(vertical="center")
    # Zahlen-/Datumsformate
    hidx = {h: i + 1 for i, h in enumerate(headers)}
    for h in date_cols:
        col = get_column_letter(hidx[h])
        for r in range(2, nrows + 1):
            ws[f"{col}{r}"].number_format = "yyyy-mm-dd"
    for h in int_cols:
        col = get_column_letter(hidx[h])
        for r in range(2, nrows + 1):
            ws[f"{col}{r}"].number_format = "#,##0"
    # ListObject
    ref = f"A1:{get_column_letter(ncols)}{nrows}"
    tbl = Table(displayName=table_name, ref=ref)
    tbl.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False)
    ws.add_table(tbl)
    ws.freeze_panes = "A2"
    # Spaltenbreiten grob
    for c in range(1, ncols + 1):
        ws.column_dimensions[get_column_letter(c)].width = max(
            12, min(34, len(str(headers[c - 1])) + 4))


# ---------------------------------------------------------------------------
# CSV-Helfer (semikolon, utf-8-sig -> deutsches Excel)
# ---------------------------------------------------------------------------

def write_csv(name, headers, rows):
    path = os.path.join(CSV_DIR, name)
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f, delimiter=";")
        w.writerow(headers)
        for r in rows:
            w.writerow([_csv_val(r.get(h)) for h in headers])


def _csv_val(v):
    if isinstance(v, dt.date):
        return v.isoformat()
    return v


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    os.makedirs(CSV_DIR, exist_ok=True)
    vol = read_source_volumes()
    fact = build_fact_actual(vol)
    plan = build_plan(fact)
    dim_date = build_dim_date()

    dim_product = [{
        "ProductKey": i + 1, "ProductName": p[1], "Category": p[2],
        "Style": p[3], "ABV": p[4], "LaunchYear": p[5],
    } for i, p in enumerate(PRODUCTS)]

    dim_customer = [{
        "CustomerKey": i + 1, "CustomerName": c[0], "CustomerType": c[1],
        "Street": c[2], "PostalCode": c[3], "City": c[4], "Region": c[5],
        "Country": c[6], "CustomerSince": c[7], "KeyAccountManager": c[8],
    } for i, c in enumerate(CUSTOMERS)]

    dim_kennzahl = [{
        "AccountKey": a[0], "AccountName": a[1], "AccountGroup": a[2],
        "Unit": a[3], "SortOrder": a[4],
    } for a in ACCOUNTS]

    # ---- DWH-Excel ----
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet("DimDate")
    write_table(ws, ["Date", "DateKey", "Year", "Quarter", "QuarterName",
                     "MonthNo", "MonthName", "MonthYear", "MonthYearSort",
                     "MonthStart", "DayOfMonth", "DayName"],
                dim_date, "tbl_DimDate",
                date_cols=("Date", "MonthStart"))

    ws = wb.create_sheet("DimProduct")
    write_table(ws, ["ProductKey", "ProductName", "Category", "Style",
                     "ABV", "LaunchYear"],
                dim_product, "tbl_DimProduct")

    ws = wb.create_sheet("DimCustomer")
    write_table(ws, ["CustomerKey", "CustomerName", "CustomerType", "Street",
                     "PostalCode", "City", "Region", "Country",
                     "CustomerSince", "KeyAccountManager"],
                dim_customer, "tbl_DimCustomer")

    ws = wb.create_sheet("DimKennzahl")
    write_table(ws, ["AccountKey", "AccountName", "AccountGroup", "Unit",
                     "SortOrder"],
                dim_kennzahl, "tbl_DimKennzahl")

    ws = wb.create_sheet("FactActual")
    write_table(ws, ["Date", "CustomerKey", "ProductKey", "AccountKey", "Value"],
                fact, "tbl_FactActual",
                date_cols=("Date",), int_cols=("Value",))

    wb.save(DWH_XLSX)

    # ---- Plan-Excel (Kreuztabelle) ----
    wb2 = openpyxl.Workbook()
    ws = wb2.active
    ws.title = "Plan"
    cat_of = {p[1]: p[2] for p in PRODUCTS}
    headers = ["Category", "Product", "Year"] + MONTHS_DE
    plan_rows = []
    for p in PRODUCT_KEY:
        for y in YEARS:
            row = {"Category": cat_of[p], "Product": p, "Year": y}
            for mi, mn in enumerate(MONTHS_DE):
                row[mn] = plan[(p, y)][mi]
            plan_rows.append(row)
    plan_rows.sort(key=lambda r: (r["Product"], r["Year"]))
    write_table(ws, headers, plan_rows, "PlanInput",
                int_cols=tuple(MONTHS_DE))
    wb2.save(PLAN_XLSX)

    # ---- CSVs fuer DB-Load ----
    write_csv("DimDate.csv", ["Date", "DateKey", "Year", "Quarter", "QuarterName",
                              "MonthNo", "MonthName", "MonthYear", "MonthYearSort",
                              "MonthStart", "DayOfMonth", "DayName"], dim_date)
    write_csv("DimProduct.csv", ["ProductKey", "ProductName", "Category", "Style",
                                 "ABV", "LaunchYear"], dim_product)
    write_csv("DimCustomer.csv", ["CustomerKey", "CustomerName", "CustomerType",
                                  "Street", "PostalCode", "City", "Region",
                                  "Country", "CustomerSince", "KeyAccountManager"],
              dim_customer)
    write_csv("DimKennzahl.csv", ["AccountKey", "AccountName", "AccountGroup",
                                  "Unit", "SortOrder"], dim_kennzahl)
    write_csv("FactActual.csv", ["Date", "CustomerKey", "ProductKey",
                                 "AccountKey", "Value"], fact)

    # ---- Report ----
    n_combos = len({(r["Date"], r["CustomerKey"], r["ProductKey"]) for r in fact})
    total_net = sum(r["Value"] for r in fact if r["AccountKey"] == 2)
    print("OK")
    print(f"  DimDate     : {len(dim_date):>6} Zeilen")
    print(f"  DimProduct  : {len(dim_product):>6} Zeilen")
    print(f"  DimCustomer : {len(dim_customer):>6} Zeilen")
    print(f"  DimKennzahl : {len(dim_kennzahl):>6} Zeilen")
    print(f"  FactActual  : {len(fact):>6} Zeilen ({n_combos} Verkaufskombis)")
    print(f"  Plan        : {len(plan_rows):>6} Zeilen (Kreuztabelle)")
    print(f"  Net Sales gesamt 2023-2026: {total_net:,.0f} EUR")
    print(f"  -> {DWH_XLSX}")
    print(f"  -> {PLAN_XLSX}")
    print(f"  -> {CSV_DIR}\\*.csv")


if __name__ == "__main__":
    main()
