#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Baut ZlatyLev_Plan.xlsx als ABSICHTLICH "gemeine" Finance-Kreuztabelle:
  - nur Plan 2026
  - geplant je Grosskunde x Produkt x Monat (Net Sales EUR)
  - Titelblock + Beschreibung, verbundene Zellen, Quartals-/Monats-Header
  - Zwischensumme unter jedem Kunden, GRAND TOTAL, Kommentar am Ende
  - EINE Seite, kein ListObject (Power Query muss es entwirren)

Liest Stamm-/Ist-Daten aus ZlatyLev_DWH.xlsx (FactActual = echte Volumina-Basis).
Datenstartzeile (fuer Power Query Table.Skip) ist fix = 10.
"""
import os
from collections import defaultdict

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DWH = os.path.join(BASE, "ZlatyLev_DWH.xlsx")
OUT = os.path.join(BASE, "ZlatyLev_Plan.xlsx")

MONTHS_DE = ["Jan", "Feb", "Mär", "Apr", "Mai", "Jun",
             "Jul", "Aug", "Sep", "Okt", "Nov", "Dez"]
GROWTH_2026 = 0.05  # Zielwachstum vs. Ist 2025

NAVY = "0F1E2E"
GOLD = "C8A24B"
BAND = "EDEAE2"      # Kundenband
SUBT = "DCD6C8"      # Zwischensumme
GTOT = "C8A24B"      # Grand total

# ---------------------------------------------------------------------------

def load_dwh():
    wb = openpyxl.load_workbook(DWH, data_only=True)
    prod = {}        # key -> (name, category, sort)
    for r in wb["DimProduct"].iter_rows(min_row=2, values_only=True):
        prod[r[0]] = {"name": r[1], "category": r[2]}
    cust = {}        # key -> dict
    for r in wb["DimCustomer"].iter_rows(min_row=2, values_only=True):
        cust[r[0]] = {"name": r[1], "type": r[2], "city": r[5]}
    # Net Sales (account 2) je (ckey, pkey, year, month)
    net = defaultdict(float)
    for d, ckey, pkey, akey, val in wb["FactActual"].iter_rows(min_row=2, values_only=True):
        if akey == 2:
            net[(ckey, pkey, d.year, d.month)] += val
    return prod, cust, net


def build_plan(prod, cust, net):
    """plan[(ckey, pkey)] = [12 Monatswerte]  (nur Kombis mit Ist 2025 > 0)"""
    # Saisonprofil je Produkt (Mittel ueber Jahre)
    annual_p = defaultdict(float)
    month_p = defaultdict(float)
    for (c, p, y, m), v in net.items():
        annual_p[(p, y)] += v
        month_p[(p, y, m)] += v
    profile = {}
    for p in prod:
        shares = [0.0] * 12
        cnt = 0
        for y in (2023, 2024, 2025, 2026):
            if annual_p[(p, y)] > 0:
                cnt += 1
                for m in range(1, 13):
                    shares[m - 1] += month_p[(p, y, m)] / annual_p[(p, y)]
        profile[p] = [s / cnt for s in shares] if cnt else [1 / 12] * 12

    # Ist 2025 je (c,p) -> Ziel 2026 -> ueber Profil verteilen
    annual_cp_2025 = defaultdict(float)
    for (c, p, y, m), v in net.items():
        if y == 2025:
            annual_cp_2025[(c, p)] += v

    plan = {}
    for (c, p), a25 in annual_cp_2025.items():
        if a25 <= 0:
            continue
        target = a25 * (1 + GROWTH_2026)
        vals = [round(target * profile[p][m] / 100) * 100 for m in range(12)]
        if sum(vals) > 0:
            plan[(c, p)] = vals
    return plan


# ---------------------------------------------------------------------------

def main():
    prod, cust, net = load_dwh()
    plan = build_plan(prod, cust, net)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Plan 2026"

    NCOL = 14  # A=Label, B..M=Monate, N=Total
    last = get_column_letter(NCOL)
    thin = Side(style="thin", color="9A9A9A")
    med = Side(style="medium", color=NAVY)
    money = '#,##0'

    def merge(r1, c1, r2, c2):
        ws.merge_cells(f"{get_column_letter(c1)}{r1}:{get_column_letter(c2)}{r2}")

    # ---- Titelblock (verbundene Zellen) ----
    merge(1, 1, 1, NCOL)
    t = ws.cell(1, 1, "ZLATÝ LEV BREWERY a.s.")
    t.font = Font(bold=True, size=18, color=NAVY)
    merge(2, 1, 2, NCOL)
    s = ws.cell(2, 1, "Prague · Smíchov   ·   Reg. No. 12345678   ·   Brewing since 1882")
    s.font = Font(italic=True, size=10, color="6B6B6B")
    # Zeile 3 leer
    merge(4, 1, 4, NCOL)
    pt = ws.cell(4, 1, "Sales Plan 2026   ·   by Key Account & Product")
    pt.font = Font(bold=True, size=14, color=NAVY)
    merge(5, 1, 5, NCOL)
    ps = ws.cell(5, 1, "Net sales plan (EUR)   ·   Finance / Controlling   ·   DRAFT v0.9 — not for distribution")
    ps.font = Font(italic=True, size=10, color=GOLD if False else "8A6D1F")
    # Zeile 6: Beschreibung (verbunden, umgebrochen)
    merge(6, 1, 6, NCOL)
    desc = ws.cell(6, 1,
        "Beschreibung: Top-down-Jahresplan je Großkunde und Produkt. Basis ist das "
        "Ist 2025 zzgl. +5 % Zielwachstum, saisonal verteilt. Werte in EUR, netto "
        "(nach Rabatt). Unter jedem Kunden steht eine Zwischensumme; ganz unten die "
        "Gesamtsumme. Achtung: Export-Zeilen sind noch in Abstimmung.")
    desc.font = Font(italic=True, size=9, color="444444")
    desc.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[6].height = 46
    # Zeile 7 leer

    # ---- Header (3 Zeilen, verbundene Zellen) ----
    hr = 8
    merge(hr, 1, hr + 2, 1)
    h0 = ws.cell(hr, 1, "Key Account / Product")
    merge(hr, 2, hr, 13)
    ws.cell(hr, 2, "Plan 2026")
    merge(hr, NCOL, hr + 2, NCOL)
    ws.cell(hr, NCOL, "Total 2026 (EUR)")
    # Quartalszeile
    for qi, c0 in enumerate((2, 5, 8, 11)):
        merge(hr + 1, c0, hr + 1, c0 + 2)
        ws.cell(hr + 1, c0, f"Q{qi + 1}")
    # Monatszeile
    for mi in range(12):
        ws.cell(hr + 2, 2 + mi, MONTHS_DE[mi])
    for rr in range(hr, hr + 3):
        for cc in range(1, NCOL + 1):
            cell = ws.cell(rr, cc)
            cell.fill = PatternFill("solid", fgColor=NAVY)
            cell.font = Font(bold=True, color="FFFFFF", size=10)
            cell.alignment = Alignment(horizontal="center", vertical="center")
    h0.alignment = Alignment(horizontal="left", vertical="center")

    # ---- Datenkörper, gruppiert nach Kunde ----
    grand = [0] * 12
    row = hr + 3  # = 11  -> Power-Query Table.Skip(10)

    # Kundenreihenfolge = CustomerKey
    for ckey in sorted(cust):
        cprods = [(p, plan[(ckey, p)]) for p in sorted(prod) if (ckey, p) in plan]
        if not cprods:
            continue
        c = cust[ckey]
        # Kunden-Kopfzeile (verbunden über alle Spalten)
        merge(row, 1, row, NCOL)
        hc = ws.cell(row, 1, f"{c['name']}   ·   {c['city']} ({c['type']})")
        hc.font = Font(bold=True, size=11, color=NAVY)
        hc.fill = PatternFill("solid", fgColor=BAND)
        hc.alignment = Alignment(vertical="center", indent=1)
        ws.row_dimensions[row].height = 18
        row += 1

        sub = [0] * 12
        for p, vals in cprods:
            ws.cell(row, 1, "      " + prod[p]["name"])
            for mi in range(12):
                cell = ws.cell(row, 2 + mi, vals[mi])
                cell.number_format = money
                sub[mi] += vals[mi]
                grand[mi] += vals[mi]
            tot = ws.cell(row, NCOL, sum(vals))
            tot.number_format = money
            tot.font = Font(color="555555")
            for cc in range(1, NCOL + 1):
                ws.cell(row, cc).border = Border(bottom=thin)
            row += 1

        # Zwischensumme je Kunde
        st = ws.cell(row, 1, f"Subtotal — {c['name']}")
        st.font = Font(bold=True, italic=True, color=NAVY)
        for mi in range(12):
            cell = ws.cell(row, 2 + mi, sub[mi])
            cell.number_format = money
            cell.font = Font(bold=True)
        ws.cell(row, NCOL, sum(sub)).number_format = money
        ws.cell(row, NCOL).font = Font(bold=True)
        for cc in range(1, NCOL + 1):
            cell = ws.cell(row, cc)
            cell.fill = PatternFill("solid", fgColor=SUBT)
            cell.border = Border(top=thin, bottom=thin)
        row += 1
        # Leerzeile zwischen Kunden
        row += 1

    # ---- GRAND TOTAL ----
    gt = ws.cell(row, 1, "GRAND TOTAL — All Key Accounts")
    gt.font = Font(bold=True, size=11, color=NAVY)
    for mi in range(12):
        cell = ws.cell(row, 2 + mi, grand[mi])
        cell.number_format = money
        cell.font = Font(bold=True, size=11, color=NAVY)
    ws.cell(row, NCOL, sum(grand)).number_format = money
    ws.cell(row, NCOL).font = Font(bold=True, size=11, color=NAVY)
    for cc in range(1, NCOL + 1):
        cell = ws.cell(row, cc)
        cell.fill = PatternFill("solid", fgColor=GTOT)
        cell.border = Border(top=med, bottom=med)
    row += 2

    # ---- Kommentar / Notes am Ende (verbunden) ----
    notes = [
        "Notes:",
        "·  Planwerte sind netto (nach Rabatt), in EUR. Mengen-/Volumenplan separat.",
        "·  Export-Kunden (Hoffmann, Tesco SK, Żabka) noch unbestätigt — Wechselkurs offen.",
        "·  Neuprodukt 'Smíchov Hazy IPA' konservativ geplant (erst seit 2025).",
        "·  Prepared by: M. Černá (Controlling).   Reviewed by: H. Svoboda.   DRAFT — do not distribute.",
    ]
    for i, line in enumerate(notes):
        rr = row + i
        merge(rr, 1, rr, NCOL)
        cell = ws.cell(rr, 1, line)
        cell.font = Font(italic=True, size=9,
                         bold=(i == 0), color="444444")
        cell.alignment = Alignment(vertical="top")

    # ---- Spaltenbreiten ----
    ws.column_dimensions["A"].width = 40
    for cc in range(2, 14):
        ws.column_dimensions[get_column_letter(cc)].width = 11
    ws.column_dimensions[last].width = 16
    ws.freeze_panes = "B11"
    ws.sheet_view.showGridLines = False

    wb.save(OUT)

    n_rows = sum(1 for ck in cust for p in prod if (ck, p) in plan)
    print("OK ->", OUT)
    print(f"  Kunden mit Plan : {len({ck for (ck, p) in plan})}")
    print(f"  Plan-Zeilen     : {n_rows} (Produkt×Kunde)")
    print(f"  GRAND TOTAL 2026: {sum(grand):,.0f} EUR")
    print("  Datenstart = Zeile 11  ->  Power Query: Table.Skip(10)")


if __name__ == "__main__":
    main()
